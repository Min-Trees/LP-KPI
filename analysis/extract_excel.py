"""Extract data from KPI Excel file for analysis."""
import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys
import io

# Force UTF-8 stdout for Windows console (cp1252 can't encode Vietnamese)
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

FILE = "Bảng chấm KPI 1.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=False)

print("=" * 80)
print(f"WORKBOOK: {FILE}")
print(f"Sheets: {wb.sheetnames}")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: '{sheet_name}'  |  Dimensions: {ws.dimensions}  |  Max row: {ws.max_row}  |  Max col: {ws.max_column}")
    print(f"{'=' * 80}")

    merged = [str(r) for r in ws.merged_cells.ranges]
    print(f"Merged ranges ({len(merged)}): {merged[:20]}{'...' if len(merged) > 20 else ''}")

    print("\n--- FULL CONTENT ---")
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            v_str = str(v)
            if len(v_str) > 120:
                v_str = v_str[:117] + "..."
            print(f"  {cell.coordinate} (row={row_idx}, col={cell.column}): {v_str!r}")

print("\n\nDONE")