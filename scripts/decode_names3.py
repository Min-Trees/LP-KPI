"""Extract ALL employee names from every sheet row-by-row."""
import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.load_workbook("Bảng chấm KPI 1.xlsx", data_only=False)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*50}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*50}")
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            col = cell.column  # 1-indexed
            val = cell.value
            if col == 1 and val and isinstance(val, (int, float)):  # col A = code
                code = int(val)
                # Find the B cell in the same row group (merged cells)
                row_cells = {c.column: c.value for c in ws[cell.row]}
                name_a = row_cells.get(2, "")   # B col
                crit_a = row_cells.get(3, "")  # C col
                print(f"  A={code}  B={name_a}  C={crit_a}")
