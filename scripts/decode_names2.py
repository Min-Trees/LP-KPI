"""Extract employee names directly from Excel."""
import openpyxl, io, sys

sys.stdout.reconfigure(encoding="utf-8")
wb = openpyxl.load_workbook("Bảng chấm KPI 1.xlsx", data_only=False)

print("=== Sheet: Manager ===")
ws = wb["Manager"]
for row in ws.iter_rows(min_row=5, max_row=7):
    # A=code, B=name, C=criterion
    a = row[0].value  # col A
    b = row[1].value  # col B
    c = row[2].value  # col C
    if a:
        print(f"  A={a} B={b} C={c}")

print("\n=== Sheet: Văn phòng + Hỗ trợ ===")
ws = wb["Văn phòng + Hỗ trợ"]
for row in ws.iter_rows(min_row=5, max_row=14):
    a = row[0].value
    b = row[1].value
    c = row[2].value
    if a:
        print(f"  A={a} B={b} C={c}")

print("\n=== Sheet: Giáo viên HS ===")
ws = wb["Giáo viên HS"]
for row in ws.iter_rows(min_row=5, max_row=12):
    a = row[0].value
    b = row[1].value
    c = row[2].value
    if a:
        print(f"  A={a} B={b} C={c}")

print("\n=== Sheet: Giáo viên ST ===")
ws = wb["Giáo viên ST"]
for row in ws.iter_rows(min_row=5, max_row=16):
    a = row[0].value
    b = row[1].value
    c = row[2].value
    if a:
        print(f"  A={a} B={b} C={c}")
